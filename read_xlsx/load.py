from openpyxl import load_workbook
import psycopg2

conn = psycopg2.connect(database="test_smile", user="postgres", password="postgres", host="127.0.0.1", port="5432")
cur=conn.cursor()

wb1 = load_workbook(filename = r"G:\PycharmProjects\read_xlsx\Google UTM Params for PAC.xlsx")
wb2 = load_workbook(filename = r"G:\PycharmProjects\read_xlsx\Bing UTM Params for PAC.xlsx")
wb3 = load_workbook(filename = r"G:\PycharmProjects\read_xlsx\Facebook-IG UTM Params for Chad.xlsx")
wb4 = load_workbook(filename = r"G:\PycharmProjects\read_xlsx\Pinterest UTM Params for PAC.xlsx")

record_list=[]


for i in wb1:
    for k in i:
        record_list.append(k)

record_list=record_list[1:]
for i in record_list:
    val = [j.value if j.value is not None else '' for j in i]

    cur.execute("INSERT INTO onetable (source,campaign,geo,prospecting_retargeting,brand_nobrand) \
          VALUES (%s,%s,%s,%s,%s)",[val[2],val[3],val[4],val[5],val[6]]);

record_list=[]

for i in wb2:
    for k in i:
        record_list.append(k)

record_list=record_list[1:]
for i in record_list:
    val = [j.value if j.value is not None else '' for j in i]

    cur.execute("INSERT INTO onetable (source,campaign,geo,prospecting_retargeting,brand_nobrand) \
          VALUES (%s,%s,%s,%s,%s)",[val[1],val[2],val[4],val[3],val[5]]);

record_list=[]

for i in wb3:
    for k in i:
        record_list.append(k)

record_list=record_list[1:]
for i in record_list:
    val = [j.value if j.value is not None else '' for j in i]
    if len(val)>7:
        break
    cur.execute("INSERT INTO onetable (source,campaign,geo,prospecting_retargeting,brand_nobrand) \
          VALUES (%s,%s,%s,%s,%s)",[val[4],val[5],val[3],val[2],'']);

record_list=[]

for i in wb4:
    for k in i:
        record_list.append(k)

record_list=record_list[1:]
for i in record_list:
    val = [j.value if j.value is not None else '' for j in i]

    cur.execute("INSERT INTO onetable (source,campaign,geo,prospecting_retargeting,brand_nobrand) \
          VALUES (%s,%s,%s,%s,%s)",[val[3],val[4],val[1],val[2],'']);

conn.commit()
conn.close()